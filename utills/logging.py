from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

def update_excel_sheet(file_path, full_report, summary):
    """
    Updates an Excel sheet with the full medical report and its summary.

    Args:
        file_path (str): Path to the Excel file.
        full_report (str): Text of the entire medical report.
        summary (str): Concise summary of the medical report.
    """
    try:
        # Load the existing workbook or create a new one if it doesn't exist
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            # Add headers if it's a new file
            sheet.append(["Full Report", "Summary"])

        # Append the new data to the sheet
        sheet.append([full_report, summary])

        # Save the workbook
        workbook.save(file_path)
        print(f"Excel sheet updated successfully at {file_path}")

    except Exception as e:
        print(f"An error occurred: {e}")




def print_summary(summary):
    print("\n\n------------------- SUMMERY OF THE REPORT ----------------------- \n\n")
    print(summary)
    print("\n\n\n")